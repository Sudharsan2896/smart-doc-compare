"""
Converters — the extra tools that live alongside the comparison engine.

    pdf_to_word            : turn a digital PDF into an editable Word (.docx)
    word_tables_to_excel   : pull every table out of a Word file into Excel (.xlsx)

Both run locally and free. Like the rest of the app, the PDF converter works on
DIGITAL PDFs (real selectable text); scanned/photographed PDFs need OCR, which is
a later roadmap item.
"""

from __future__ import annotations

import io
import os
import tempfile


def pdf_to_word(pdf_bytes: bytes) -> bytes:
    """Convert a digital PDF into a Word .docx and return the new file's bytes."""
    from pdf2docx import Converter

    # pdf2docx works with files on disk, so we use a short-lived temp folder that
    # is deleted automatically when we're done.
    with tempfile.TemporaryDirectory() as tmp:
        pdf_path = os.path.join(tmp, "input.pdf")
        docx_path = os.path.join(tmp, "output.docx")

        with open(pdf_path, "wb") as f:
            f.write(pdf_bytes)

        converter = Converter(pdf_path)
        try:
            converter.convert(docx_path)  # convert all pages
        finally:
            converter.close()

        with open(docx_path, "rb") as f:
            return f.read()


def word_tables_to_excel(docx_bytes: bytes) -> tuple[bytes, int]:
    """
    Extract every table from a Word .docx into an Excel workbook.

    Each Word table becomes its own sheet ("Table 1", "Table 2", …). Returns the
    Excel file's bytes plus the number of tables found (0 if the document has none).
    """
    import docx  # python-docx
    from openpyxl import Workbook

    document = docx.Document(io.BytesIO(docx_bytes))
    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet; we add our own

    count = 0
    for i, table in enumerate(document.tables, start=1):
        ws = wb.create_sheet(title=f"Table {i}")  # sheet names stay well under 31 chars
        for row in table.rows:
            ws.append([cell.text.strip() for cell in row.cells])
        count += 1

    if count == 0:
        ws = wb.create_sheet(title="No tables found")
        ws.append(["No tables were found in this Word document."])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), count
